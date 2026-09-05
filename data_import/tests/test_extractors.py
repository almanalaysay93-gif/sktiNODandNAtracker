import os
import tempfile
import unittest
import pandas as pd
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from data_import.extractors.excel_extractor import extract_excel_sheets, detect_header_row
from data_import.extractors.pdf_extractor import extract_pdf_document


class TestExtractors(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_excel_header_detection_and_unmerging(self):
        wb_path = os.path.join(self.temp_dir.name, "hospital_roster.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "StaffRoster"

        # Row 1: Hospital title banner
        ws.cell(row=1, column=1, value="SOUTHERN PHILIPPINES MEDICAL CENTER")
        ws.cell(row=2, column=1, value="NURSING SERVICE ROSTER 2026")
        # Row 3: Empty row
        # Row 4: Merged category header
        ws.cell(row=4, column=1, value="STAFF DETAILS")
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        ws.cell(row=4, column=3, value="UNIT DETAILS")
        ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4)

        # Row 5: Actual table header
        headers = ["Emp ID", "Full Name", "Area", "Position"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=5, column=col_idx, value=h)

        # Row 6-7: Data
        ws.cell(row=6, column=1, value="RN-101")
        ws.cell(row=6, column=2, value="Dela Cruz, Juan M.")
        ws.cell(row=6, column=3, value="ICU")
        ws.cell(row=6, column=4, value="Staff Nurse")

        ws.cell(row=7, column=1, value="RN-102")
        ws.cell(row=7, column=2, value="Reyes, Maria S.")
        ws.cell(row=7, column=3, value="ER")
        ws.cell(row=7, column=4, value="Head Nurse")

        wb.save(wb_path)

        sheets = extract_excel_sheets(wb_path)
        self.assertIn("StaffRoster", sheets)
        df = sheets["StaffRoster"]

        self.assertIsInstance(df, pd.DataFrame)
        self.assertEqual(len(df), 2)
        # Check column names match
        cols = list(df.columns)
        self.assertIn("Emp ID", cols)
        self.assertIn("Full Name", cols)
        self.assertIn("Area", cols)
        self.assertIn("Position", cols)

        # Check values
        first_row = df.iloc[0]
        self.assertEqual(str(first_row["Emp ID"]), "RN-101")
        self.assertEqual(str(first_row["Full Name"]), "Dela Cruz, Juan M.")
        self.assertEqual(str(first_row["Area"]), "ICU")

    def test_pdf_extraction_and_scanned_page_detection(self):
        pdf_path = os.path.join(self.temp_dir.name, "seminar_attendance.pdf")
        c = canvas.Canvas(pdf_path, pagesize=letter)

        # Page 1: Text document with headers and table rows
        c.setFont("Helvetica-Bold", 16)
        c.drawString(72, 750, "BASIC LIFE SUPPORT TRAINING")
        c.setFont("Helvetica", 12)
        c.drawString(72, 720, "Date: 2026-08-15")
        c.drawString(72, 700, "Provider: Heart Association")
        c.drawString(72, 670, "Employee ID | Full Name | Role | Hours")
        c.drawString(72, 650, "RN-201 | Santos, Ana | Participant | 8")
        c.drawString(72, 630, "RN-202 | Garcia, Jose | Speaker | 8")
        c.showPage()

        # Page 2: Blank / simulated scanned image page (0 text characters)
        c.setFont("Helvetica", 10)
        # Draw nothing or just an empty line
        c.showPage()
        c.save()

        result = extract_pdf_document(pdf_path)
        self.assertEqual(result["page_count"], 2)
        self.assertTrue(result["has_scanned_pages"])
        self.assertIn(2, result["scanned_page_numbers"])

        md = result["markdown"]
        self.assertIn("# BASIC LIFE SUPPORT TRAINING", md)
        self.assertIn("**Date**: 2026-08-15", md)
        # Check that table format is present
        self.assertIn("| Employee ID | Full Name | Role | Hours |", md)
        self.assertIn("| RN-201 | Santos, Ana | Participant | 8 |", md)


if __name__ == "__main__":
    unittest.main()

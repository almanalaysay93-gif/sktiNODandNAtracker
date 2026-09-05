import os
from typing import Any, Dict, List, Optional
import openpyxl
import pandas as pd


def unmerge_and_fill(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Un-merge cells and populate all cells in the span with top-left value."""
    merged_ranges = list(ws.merged_cells.ranges)
    for cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = cell_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(cell_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_val)


def detect_header_row(rows: List[List[Any]], max_scan: int = 15) -> int:
    """
    Score the first max_scan rows to find the most likely table header.
    Returns 0-based row index.
    """
    best_idx = 0
    best_score = -1.0

    common_header_tokens = {
        "id", "empid", "emp", "employee", "name", "fullname", "firstname",
        "lastname", "position", "area", "ward", "station", "unit", "status",
        "role", "title", "training", "seminar", "date", "hours", "cpd",
        "license", "prc", "category", "contact", "hired", "department"
    }

    scan_limit = min(len(rows), max_scan)
    for idx in range(scan_limit):
        row = rows[idx]
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if not non_empty:
            continue

        # Criteria 1: Number of non-empty columns (headers usually span multiple columns)
        count_score = len(non_empty) * 1.5

        # Criteria 2: Matching known header tokens
        token_matches = 0
        all_short_strings = True
        for cell in non_empty:
            s = str(cell).strip().lower()
            # If any cell is a long sentence (> 60 chars), likely title or notice, not header
            if len(s) > 60:
                all_short_strings = False
            clean_words = "".join(ch if ch.isalnum() else " " for ch in s).split()
            if any(w in common_header_tokens for w in clean_words):
                token_matches += 1

        token_score = token_matches * 5.0
        string_bonus = 2.0 if all_short_strings and len(non_empty) >= 2 else -5.0

        # Check subsequent row: if subsequent row has data, this row is likely header
        has_next_row_data = False
        if idx + 1 < len(rows):
            next_row = [c for c in rows[idx + 1] if c is not None and str(c).strip() != ""]
            if len(next_row) >= len(non_empty) - 1:
                has_next_row_data = True
        data_bonus = 3.0 if has_next_row_data else 0.0

        total_score = count_score + token_score + string_bonus + data_bonus
        if total_score > best_score:
            best_score = total_score
            best_idx = idx

    return best_idx


def extract_excel_sheets(filepath: str) -> Dict[str, pd.DataFrame]:
    """
    Load an Excel workbook, unmerge cells, locate table headers, and return
    a dictionary of {sheet_name: clean_dataframe}.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result: Dict[str, pd.DataFrame] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        unmerge_and_fill(ws)

        raw_rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))

        if not raw_rows:
            continue

        # Filter trailing completely empty rows
        while raw_rows and not any(c is not None and str(c).strip() != "" for c in raw_rows[-1]):
            raw_rows.pop()

        if not raw_rows:
            continue

        header_idx = detect_header_row(raw_rows)
        raw_headers = raw_rows[header_idx]

        # Clean header names and resolve duplicates or blank names
        headers: List[str] = []
        seen_headers: Dict[str, int] = {}
        for col_idx, h in enumerate(raw_headers):
            h_str = str(h).strip() if h is not None and str(h).strip() != "" else f"Column_{col_idx + 1}"
            if h_str in seen_headers:
                seen_headers[h_str] += 1
                headers.append(f"{h_str}_{seen_headers[h_str]}")
            else:
                seen_headers[h_str] = 1
                headers.append(h_str)

        data_rows = raw_rows[header_idx + 1:]
        if not data_rows:
            df = pd.DataFrame(columns=headers)
        else:
            # Pad or trim data rows to match headers length
            normalized_rows = []
            for r in data_rows:
                # Skip rows that are completely empty
                if not any(c is not None and str(c).strip() != "" for c in r):
                    continue
                padded = r[:len(headers)] + [None] * max(0, len(headers) - len(r))
                normalized_rows.append(padded)
            df = pd.DataFrame(normalized_rows, columns=headers)

        # Drop empty columns
        df = df.dropna(how="all", axis=1)
        result[sheet_name] = df

    return result

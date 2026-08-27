import openpyxl
import json
import re
from datetime import datetime, date

EXCEL_PATH = r"C:\Users\Admin\Downloads\NN LDI DATABASE SUMMARY.xlsx"
OUTPUT_JSON = r"E:\ai\claude\ALAi\skti-nursetrack\server\data\seedData.json"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try YYYY-MM-DD
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # Try MM/DD/YYYY or DD/MM/YYYY
    m = re.search(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", s)
    if m:
        p1, p2, p3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if p3 < 100:
            p3 += 2000
        # If p1 > 12, then p1 is day, p2 is month
        if p1 > 12:
            return f"{p3:04d}-{p2:02d}-{p1:02d}"
        else:
            return f"{p3:04d}-{p1:02d}-{p2:02d}"
    return None

def parse_name(raw_name):
    raw = clean_str(raw_name)
    parts = raw.split(",", 1)
    if len(parts) == 2:
        last = parts[0].strip()
        first_rest = parts[1].strip()
        suffix = None
        for s in ["Jr.", "Jr", "Sr.", "Sr", "III", "II", "IV"]:
            if first_rest.endswith(" " + s) or first_rest.endswith("," + s):
                suffix = s
                first_rest = first_rest[:-len(s)].strip().rstrip(",")
                break
        tokens = first_rest.split()
        if len(tokens) == 0:
            first = ""
            middle = None
        elif len(tokens) == 1:
            first = tokens[0]
            middle = None
        else:
            if len(tokens[-1].rstrip(".")) <= 2:
                first = " ".join(tokens[:-1])
                middle = tokens[-1]
            else:
                first = " ".join(tokens[:-1])
                middle = tokens[-1]
        return {
            "fullName": raw,
            "lastName": last.title(),
            "firstName": first.title(),
            "middleName": middle.title() if middle else None,
            "suffix": suffix
        }
    else:
        tokens = raw.split()
        if len(tokens) >= 2:
            # Check if name is duplicated string
            half = len(tokens) // 2
            if tokens[:half] == tokens[half:]:
                tokens = tokens[:half]
            # Staff rosters use LAST First Middle format when a comma is
            # accidentally omitted (for example, "SOTTO Mary Grace S.").
            # A trailing initial distinguishes this from ordinary First Last.
            if len(tokens) >= 3 and len(tokens[-1].rstrip(".")) <= 2:
                return {
                    "fullName": raw,
                    "lastName": tokens[0].title(),
                    "firstName": " ".join(tokens[1:-1]).title(),
                    "middleName": tokens[-1].title(),
                    "suffix": None
                }
            last = tokens[-1]
            first = " ".join(tokens[:-1])
            return {
                "fullName": raw,
                "lastName": last.title(),
                "firstName": first.title(),
                "middleName": None,
                "suffix": None
            }
        return {
            "fullName": raw,
            "lastName": raw.title(),
            "firstName": "",
            "middleName": None,
            "suffix": None
        }

# Define Default Areas
AREAS = [
    {"code": "NEPHRO-OFFICE", "name": "Nephrology Office", "description": "Nephrology Nursing Office & Administrative Center", "sortOrder": 1},
    {"code": "PD", "name": "Peritoneal Dialysis", "description": "Peritoneal Dialysis Unit & Outpatient CAPD/APD", "sortOrder": 2},
    {"code": "OTSU-SHARE", "name": "OTSU / SHARE", "description": "Organ Transplant Specialty Unit & SHARE Programs", "sortOrder": 3},
    {"code": "RDU-MAIN", "name": "RDU Main", "description": "Renal Dialysis Unit - Main Building (Station 1-28)", "sortOrder": 4},
    {"code": "RDU-ANNEX", "name": "RDU Annex", "description": "Renal Dialysis Unit - Annex Center", "sortOrder": 5},
    {"code": "SKTI-WARD", "name": "SKTI Service Ward", "description": "Southern Philippines Kidney Transplant Institute - Inpatient Ward", "sortOrder": 6},
    {"code": "SKTI-PAY", "name": "SKTI Payward", "description": "SKTI Pay Patients Inpatient Unit", "sortOrder": 7},
    {"code": "SKTI-ICU", "name": "SKTI ICU", "description": "SKTI Intensive Care Unit", "sortOrder": 8},
    {"code": "TRIAGE", "name": "Triage & Receiving", "description": "Nephrology Triage and Outpatient Receiving", "sortOrder": 9},
]

area_map_by_keyword = {
    "nephrology nursing office": "NEPHRO-OFFICE",
    "nephrology office": "NEPHRO-OFFICE",
    "peritoneal dialysis": "PD",
    "peritoneal dialysis nurses": "PD",
    "share nurses": "OTSU-SHARE",
    "otsu nurses": "OTSU-SHARE",
    "otsu/share": "OTSU-SHARE",
    "du main": "RDU-MAIN",
    "du main nurses": "RDU-MAIN",
    "rdu main": "RDU-MAIN",
    "du annex": "RDU-ANNEX",
    "du annex nurses": "RDU-ANNEX",
    "rdu annex": "RDU-ANNEX",
    "skti service ward": "SKTI-WARD",
    "skti ward": "SKTI-WARD",
    "skti payward": "SKTI-PAY",
    "skti pay": "SKTI-PAY",
    "skti icu": "SKTI-ICU",
    "hd tech": "RDU-MAIN",
    "triage": "TRIAGE",
    "skti": "SKTI-WARD",
}

# Collect Staff
staff_dict = {} # key by normalized name
training_catalog_dict = {} # key by normalized name
events_list = []
attendance_records = []

# Known source spelling variants in the quarterly ledgers. Values match the
# canonical roster key generated by parse_name().
STAFF_NAME_ALIASES = {
    "CONCEPTION, DANDREB": "CONCEPCION, DANREB",
}

def resolve_staff(norm_name):
    """Resolve a ledger name to exactly one roster entry."""
    canonical = STAFF_NAME_ALIASES.get(norm_name, norm_name)
    if canonical in staff_dict:
        return canonical, staff_dict[canonical]

    last_name, _, first_name = canonical.partition(",")
    last_name = last_name.strip()
    first_name = first_name.strip()
    candidates = [
        (key, person)
        for key, person in staff_dict.items()
        if key.partition(",")[0].strip() == last_name
    ]
    if len(candidates) == 1:
        return candidates[0]

    prefix_matches = [
        (key, person)
        for key, person in candidates
        if key.partition(",")[2].strip().startswith(first_name)
        or first_name.startswith(key.partition(",")[2].strip())
    ]
    if len(prefix_matches) == 1:
        return prefix_matches[0]

    raise ValueError(f"Could not resolve staff name uniquely: {norm_name}")

# 1. Parse NURSES sheet
sheet_nurses = wb["NURSES"]
current_area_code = "NEPHRO-OFFICE"

# Row 4 has training headers
nurse_training_headers = {}
for c in range(8, sheet_nurses.max_column + 1):
    th = sheet_nurses.cell(4, c).value
    if th and clean_str(th):
        th_clean = clean_str(th)
        nurse_training_headers[c] = th_clean
        if th_clean.lower() not in training_catalog_dict:
            training_catalog_dict[th_clean.lower()] = {
                "name": th_clean,
                "category": "Mandatory / Clinical",
                "kind": "Training" if "training" in th_clean.lower() or "first aid" in th_clean.lower() or "iv" in th_clean.lower() else "Seminar",
                "renewalRequired": any(k in th_clean.lower() for k in ["bls", "acls", "first aid", "iv", "sfat"]),
                "defaultValidityMonths": 24 if any(k in th_clean.lower() for k in ["bls", "acls", "first aid"]) else None
            }

for r in range(6, sheet_nurses.max_row + 1):
    col1 = sheet_nurses.cell(r, 1).value
    col2 = sheet_nurses.cell(r, 2).value
    if not col2 or not clean_str(col2):
        continue
    
    col2_str = clean_str(col2)
    # Check if section header
    if col1 is None or clean_str(col1) == "":
        col2_lower = col2_str.lower()
        if "legend" in col2_lower or "2023" in col2_lower:
            break
        for kw, code in area_map_by_keyword.items():
            if kw in col2_lower:
                current_area_code = code
                break
        continue
    
    col2_lower = col2_str.lower()
    if any(k in col2_lower for k in ["legend", "near expiry", "not yet employed", "for update", "nurses", "nursing attendants", "summary", "total"]):
        continue
    # It is a nurse row
    name_info = parse_name(col2_str)
    email = clean_str(sheet_nurses.cell(r, 3).value) or None
    raw_license = sheet_nurses.cell(r, 4).value
    license_no = ""
    if raw_license is not None:
        if isinstance(raw_license, float):
            license_no = str(int(raw_license)) if raw_license == int(raw_license) else str(raw_license)
        else:
            license_no = str(raw_license).strip()
    
    expiry = parse_date(sheet_nurses.cell(r, 5).value)
    hist_2023 = clean_str(sheet_nurses.cell(r, 6).value)
    hist_2024 = clean_str(sheet_nurses.cell(r, 7).value)
    
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    emp_id = f"RN-{len(staff_dict)+1:03d}"
    
    nurse_obj = {
        "employeeId": emp_id,
        "nameInfo": name_info,
        "email": email,
        "staffType": "Registered Nurse",
        "position": "Staff Nurse II" if current_area_code != "NEPHRO-OFFICE" else "Nurse Supervisor",
        "employmentStatus": "Active",
        "currentAreaCode": current_area_code,
        "licenseNumber": license_no,
        "licenseExpiry": expiry,
        "historyNotes": f"2023 & Earlier: {hist_2023} | 2024: {hist_2024}".strip(" |"),
        "matrixTrainings": {}
    }
    
    # Read training cells
    for col_idx, th_name in nurse_training_headers.items():
        cell_val = sheet_nurses.cell(r, col_idx).value
        if cell_val is not None and clean_str(cell_val):
            val_str = clean_str(cell_val)
            nurse_obj["matrixTrainings"][th_name] = val_str
            
    staff_dict[norm_name] = nurse_obj

print(f"Loaded {len(staff_dict)} nurses from NURSES sheet.")

# 2. Parse NURSING ATTENDANTS sheet
sheet_na = wb["NURSING ATTENDANTS"]
na_training_headers = {}
for c in range(6, sheet_na.max_column + 1):
    th = sheet_na.cell(3, c).value
    if th and clean_str(th):
        th_clean = clean_str(th)
        na_training_headers[c] = th_clean
        if th_clean.lower() not in training_catalog_dict:
            training_catalog_dict[th_clean.lower()] = {
                "name": th_clean,
                "category": "Nursing Attendant / General",
                "kind": "Training" if "training" in th_clean.lower() or "drill" in th_clean.lower() else "Seminar",
                "renewalRequired": any(k in th_clean.lower() for k in ["bls", "fire drill", "first aid"]),
                "defaultValidityMonths": 24 if "bls" in th_clean.lower() else None
            }

for r in range(5, sheet_na.max_row + 1):
    col1 = sheet_na.cell(r, 1).value
    col2 = sheet_na.cell(r, 2).value
    if not col2 or not clean_str(col2):
        continue
    col2_str = clean_str(col2)
    if any(k in col2_str.lower() for k in ["legend", "near expiry", "not yet employed", "for update", "from netu", "2023", "2024", "2025", "total", "summary", "nurses", "nursing attendants"]):
        continue
    name_info = parse_name(col2_str)
    
    raw_lic = sheet_na.cell(r, 3).value
    lic_no = ""
    if raw_lic is not None:
        if isinstance(raw_lic, float):
            lic_no = str(int(raw_lic)) if raw_lic == int(raw_lic) else str(raw_lic)
        else:
            lic_no = str(raw_lic).strip().rstrip("/")
            
    expiry = parse_date(sheet_na.cell(r, 4).value)
    email = clean_str(sheet_na.cell(r, 5).value) or None
    
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    emp_id = f"NA-{len([s for s in staff_dict.values() if s['staffType'] == 'Nursing Attendant']) + 1:03d}"
    
    # Check if this NA has area hint from List of All Nursing Attendants
    na_area = "RDU-MAIN"
    
    na_obj = {
        "employeeId": emp_id,
        "nameInfo": name_info,
        "email": email,
        "staffType": "Nursing Attendant",
        "position": "Nursing Attendant I",
        "employmentStatus": "Active",
        "currentAreaCode": na_area,
        "licenseNumber": lic_no,
        "licenseExpiry": expiry,
        "historyNotes": "",
        "matrixTrainings": {}
    }
    
    for col_idx, th_name in na_training_headers.items():
        cell_val = sheet_na.cell(r, col_idx).value
        if cell_val is not None and clean_str(cell_val):
            na_obj["matrixTrainings"][th_name] = clean_str(cell_val)
            
    staff_dict[norm_name] = na_obj

print(f"Total staff after NURSING ATTENDANTS: {len(staff_dict)}")

# 3. Parse RotationResignees sheet
sheet_rr = wb["RotationResignees"]
for r in range(2, sheet_rr.max_row + 1):
    col1 = clean_str(sheet_rr.cell(r, 1).value).upper() # ROTATION or RESIGNEE
    col2 = sheet_rr.cell(r, 2).value # ID
    col3 = clean_str(sheet_rr.cell(r, 3).value) # Name
    if not col3:
        continue
    name_info = parse_name(col3)
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    
    email = clean_str(sheet_rr.cell(r, 4).value) or None
    raw_lic = sheet_rr.cell(r, 5).value
    lic_no = ""
    if raw_lic is not None:
        if isinstance(raw_lic, float):
            lic_no = str(int(raw_lic)) if raw_lic == int(raw_lic) else str(raw_lic)
        else:
            lic_no = str(raw_lic).strip()
    expiry = parse_date(sheet_rr.cell(r, 6).value)
    status = "Rotated" if "ROTATION" in col1 else "Resigned"
    
    if norm_name in staff_dict:
        staff_dict[norm_name]["employmentStatus"] = status
        if email and not staff_dict[norm_name]["email"]:
            staff_dict[norm_name]["email"] = email
        if lic_no and not staff_dict[norm_name]["licenseNumber"]:
            staff_dict[norm_name]["licenseNumber"] = lic_no
        if expiry and not staff_dict[norm_name]["licenseExpiry"]:
            staff_dict[norm_name]["licenseExpiry"] = expiry
    else:
        emp_id = f"RN-{len(staff_dict)+1:03d}"
        staff_dict[norm_name] = {
            "employeeId": emp_id,
            "nameInfo": name_info,
            "email": email,
            "staffType": "Registered Nurse",
            "position": "Staff Nurse",
            "employmentStatus": status,
            "currentAreaCode": "RDU-MAIN",
            "licenseNumber": lic_no,
            "licenseExpiry": expiry,
            "historyNotes": "",
            "matrixTrainings": {}
        }

print(f"Total staff after RotationResignees: {len(staff_dict)}")

# 4. Parse List of All Nursing Attendants for Area notes
sheet_lana = wb["List of All Nursing Attendants "]
for r in range(1, sheet_lana.max_row + 1):
    col1 = clean_str(sheet_lana.cell(r, 1).value)
    col5 = clean_str(sheet_lana.cell(r, 5).value) # e.g. '28 - hd tech', '1- triage', '10- SKTI'
    if not col1:
        continue
    name_info = parse_name(col1)
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    if norm_name in staff_dict and col5:
        col5_lower = col5.lower()
        if "triage" in col5_lower:
            staff_dict[norm_name]["currentAreaCode"] = "TRIAGE"
        elif "skti" in col5_lower:
            staff_dict[norm_name]["currentAreaCode"] = "SKTI-WARD"
        elif "hd tech" in col5_lower or "hd" in col5_lower:
            staff_dict[norm_name]["currentAreaCode"] = "RDU-MAIN"

# 5. Parse 1ST QUARTER SUMMARY
# Format: Row 6: [No., NAME, TITLE OF SEMINAR/TRAINING, DATE CONDUCTED]
# Some cells have multiple titles separated by newlines and matching dates
sheet_q1 = wb["1ST QUARTER SUMMARY"]
q1_entries = []

for r in range(7, sheet_q1.max_row + 1):
    col2 = clean_str(sheet_q1.cell(r, 2).value)
    col3 = sheet_q1.cell(r, 3).value
    col4 = sheet_q1.cell(r, 4).value
    if not col2 or not col3:
        continue
    
    name_info = parse_name(col2)
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    norm_name, matched_staff = resolve_staff(norm_name)
    
    # Split multiline titles and dates
    titles = [t.strip() for t in str(col3).split("\n") if t.strip()]
    dates = [col4] if isinstance(col4, (datetime, date)) else ([d.strip() for d in str(col4).split("\n") if d.strip()] if col4 else [])
    
    for idx, title in enumerate(titles):
        raw_date = dates[idx] if idx < len(dates) else (dates[0] if dates else "Q1 2026")
        d_str = clean_str(raw_date)
        
        # Try to parse start/end date from date string
        # e.g. 'March 18, 2026 (9am-12nn)', 'Jan 14, 2026, 8:00 PM', '02/19,20,23; 8am-5pm', '03/16-17/26'
        start_d = None
        end_d = None
        
        if isinstance(raw_date, (datetime, date)):
            start_d = raw_date.strftime("%Y-%m-%d")
            end_d = start_d
        else:
            # Match Month Name Day, 2026
            m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{1,2})(?:-(\d{1,2}))?,?\s+(\d{4})", d_str, re.IGNORECASE)
        if not isinstance(raw_date, (datetime, date)) and m:
            month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
            mon = month_map[m.group(1).lower()[:3]]
            d1 = int(m.group(2))
            d2 = int(m.group(3)) if m.group(3) else d1
            yr = int(m.group(4))
            start_d = f"{yr:04d}-{mon:02d}-{d1:02d}"
            end_d = f"{yr:04d}-{mon:02d}-{d2:02d}"
        elif not isinstance(raw_date, (datetime, date)):
            # Match MM/DD/YY
            m2 = re.search(r"(\d{1,2})[/.-](\d{1,2})(?:-(\d{1,2}))?[/.-](\d{2,4})", d_str)
            if m2:
                mon = int(m2.group(1))
                d1 = int(m2.group(2))
                d2 = int(m2.group(3)) if m2.group(3) else d1
                yr = int(m2.group(4))
                if yr < 100: yr += 2000
                start_d = f"{yr:04d}-{mon:02d}-{d1:02d}"
                end_d = f"{yr:04d}-{mon:02d}-{d2:02d}"
            else:
                # Default Q1 date
                start_d = "2026-02-15"
                end_d = "2026-02-15"
                
        if title.lower() not in training_catalog_dict:
            training_catalog_dict[title.lower()] = {
                "name": title,
                "category": "LDI / Professional Development",
                "kind": "LDI",
                "renewalRequired": False,
                "defaultValidityMonths": None
            }

        q1_entries.append({
            "quarter": 1,
            "staffName": col2,
            "normName": norm_name,
            "employeeId": matched_staff["employeeId"],
            "title": title,
            "rawDate": d_str,
            "startDate": start_d,
            "endDate": end_d,
            "provider": "SPMC NETU / PETD",
            "role": "Participant"
        })

print(f"Extracted {len(q1_entries)} individual attendance records from 1ST QUARTER SUMMARY.")

# 6. Parse 2ND QUARTER SUMMARY
# Format: Row 6: ['', NAME OF STAFF, COMPLETE TITLE OF SEMINAR/WEBINAR/ TRAINING ATTENDED, DATE & TIME, CONDUCTED BY]
sheet_q2 = wb["2ND QUARTER SUMMARY"]
q2_entries = []

for r in range(7, sheet_q2.max_row + 1):
    col2 = clean_str(sheet_q2.cell(r, 2).value)
    col3 = sheet_q2.cell(r, 3).value
    col4 = sheet_q2.cell(r, 4).value
    col5 = clean_str(sheet_q2.cell(r, 5).value) or "SPMC NETU"
    if not col2 or not col3:
        continue
        
    title = clean_str(col3)
    name_info = parse_name(col2)
    norm_name = name_info["lastName"].upper() + ", " + name_info["firstName"].upper()
    norm_name, matched_staff = resolve_staff(norm_name)
    
    # Check if speaker/resource speaker
    role = "Participant"
    if "RESOURCE SPEAKER" in title.upper() or "SPEAKER" in title.upper():
        role = "Speaker"
        # clean title
        title = re.sub(r"\(RESOURCE SPEAKER\)", "", title, flags=re.IGNORECASE).strip()
    elif "FACILITATOR" in title.upper():
        role = "Facilitator"
        title = re.sub(r"\(FACILITATOR\)", "", title, flags=re.IGNORECASE).strip()
    elif "PRECEPTOR" in title.upper():
        role = "Preceptor"
        
    d_str = clean_str(col4)
    start_d = None
    end_d = None
    
    if isinstance(col4, (datetime, date)):
        start_d = col4.strftime("%Y-%m-%d")
        end_d = start_d
    else:
        m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{1,2})(?:-(\d{1,2}))?,?\s+(\d{4})", d_str, re.IGNORECASE)
        if m:
            month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
            mon = month_map[m.group(1).lower()[:3]]
            d1 = int(m.group(2))
            d2 = int(m.group(3)) if m.group(3) else d1
            yr = int(m.group(4))
            start_d = f"{yr:04d}-{mon:02d}-{d1:02d}"
            end_d = f"{yr:04d}-{mon:02d}-{d2:02d}"
        else:
            m2 = re.search(r"(\d{1,2})[/.-](\d{1,2})(?:-(\d{1,2}))?[/.-](\d{2,4})", d_str)
            if m2:
                mon = int(m2.group(1))
                d1 = int(m2.group(2))
                d2 = int(m2.group(3)) if m2.group(3) else d1
                yr = int(m2.group(4))
                if yr < 100: yr += 2000
                start_d = f"{yr:04d}-{mon:02d}-{d1:02d}"
                end_d = f"{yr:04d}-{mon:02d}-{d2:02d}"
            else:
                start_d = "2026-05-15"
                end_d = "2026-05-15"
                
    if title.lower() not in training_catalog_dict:
        training_catalog_dict[title.lower()] = {
            "name": title,
            "category": "LDI / Professional Development",
            "kind": "LDI",
            "renewalRequired": False,
            "defaultValidityMonths": None
        }
        
    q2_entries.append({
        "quarter": 2,
        "staffName": col2,
        "normName": norm_name,
        "employeeId": matched_staff["employeeId"],
        "title": title,
        "rawDate": d_str,
        "startDate": start_d,
        "endDate": end_d,
        "provider": col5,
        "role": role
    })

print(f"Extracted {len(q2_entries)} individual attendance records from 2ND QUARTER SUMMARY.")

# 7. Group Events by (Title, StartDate, EndDate, Provider)
event_groups = {}
for entry in q1_entries + q2_entries:
    key = (entry["title"], entry["startDate"], entry["endDate"], entry["provider"])
    if key not in event_groups:
        event_groups[key] = {
            "title": entry["title"],
            "startDate": entry["startDate"],
            "endDate": entry["endDate"],
            "provider": entry["provider"],
            "venue": "SPMC Training Hall / Virtual",
            "attendees": []
        }
    event_groups[key]["attendees"].append({
        "staffName": entry["staffName"],
        "normName": entry["normName"],
        "employeeId": entry["employeeId"],
        "role": entry["role"],
        "completionDate": entry["endDate"]
    })

print(f"Consolidated into {len(event_groups)} unique seminar events.")

# Structure the final output JSON
output_data = {
    "areas": AREAS,
    "trainingCatalog": list(training_catalog_dict.values()),
    "staff": list(staff_dict.values()),
    "events": list(event_groups.values()),
    "q1Count": len(q1_entries),
    "q2Count": len(q2_entries),
    "totalStaff": len(staff_dict),
    "totalCatalog": len(training_catalog_dict)
}

import os
os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(output_data, f, indent=2)

print(f"Successfully generated {OUTPUT_JSON} with {len(output_data['staff'])} staff, {len(output_data['trainingCatalog'])} training catalog items, and {len(output_data['events'])} events!")
